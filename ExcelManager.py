###########################################################
#
#   Excelファイルを管理するクラス
#
#       指定されたファイルを読み込んで、
#
###########################################################
import openpyxl
import sys
import os
import shutil


def make_save_filename(file_name):
    """
    保存するファイル名を作成する
    指定したファイル名の後ろに"_番号"をつけたファイル名を作る
    但し、既に_番号をつけたファイル名があれば、ユニークな名前になるまで番号をインクリメントする
    :param file_name:
    :return:候補ファイル名
    :note:10000までやっても重複する場合は、exit(1)で抜ける
    """
    basename_without_ext = os.path.splitext(os.path.basename(file_name))[0]  # 拡張子を除いた名前
    for count in range(0, 10000):
        basename = basename_without_ext + "_%d" % count  # "_番号"をつける
        dir_name = os.path.dirname(file_name)  # ディレクトリパス名
        ext_name = os.path.splitext(file_name)[1][1:]  # 拡張子
        if ext_name != '':  # 拡張子があれば、間に'.'をつける
            basename = basename + '.' + ext_name
        filename_candidacy = os.path.join(dir_name, basename)  # パス名を再構築
        if not os.path.exists(filename_candidacy):  # 生成したファイル名が存在しなければ、返す
            return filename_candidacy  # 存在しなければその名前に決定

    print("%sからファイル名を作れませんでした" % file_name)  # 散々探して諦めた
    sys.exit(1)


class ExcelManager:
    def __init__(self, file_name):
        """
        指定されたExcelファイルを管理する
        :param file_name:ファイル名
        """
        save_file_name = make_save_filename(file_name)  # 生成するファイル名

        try:
            shutil.copy(file_name, save_file_name)  # ファイルを複写
        except FileNotFoundError as err:
            print(type(err))  # ファイルが見つからなかった
            print(err)
            print("ファイル%sがありません" % file_name)
            sys.exit(1)

        except OSError as err:
            print(type(err))  # その他、コピー出来ない事情があった
            print(err)
            print("%sをファイル%sに複製出来ませんでした" % (file_name, save_file_name))
            sys.exit(1)

        self.file_name = save_file_name  # ファイル名確定
        try:
            # Excel Bookを開く
            self.work_book = openpyxl.load_workbook(self.file_name)
        except OSError as err:
            print(type(err))  # ファイルを開けなかった
            print(err)
            print("ファイル%sがオープン出来ませんでした" % self.file_name)
            sys.exit(1)

    def close(self):
        """
        ファイルを閉じる
        """
        try:
            self.work_book.save(self.file_name)
        except OSError as err:
            print(type(err))  # 保存に失敗した
            print(err)
            print("ファイル%sが保存出来ませんでした" % self.file_name)
            sys.exit(1)

        self.work_book.close()

    def get_sheet(self, sheet_name):
        """
        シートオブジェクトを得る
        :param sheet_name:シート名
        :return:シートオブジェクト(無い時はNone)
        """
        if sheet_name in self.work_book.sheetnames:
            work_sheet = self.work_book[sheet_name]
            return work_sheet
        else:
            return None

    def get_column(self, sheet_obj, column_name, column_name_line):
        """
        指定した行番号をラベル名の行とし、column_nameで指定された文字列をラベル名とするカラムを返す
        :param sheet_obj:チェックするシート
        :param column_name:検索するラベル名
        :param column_name_line:検索する行
        :return:見つけたカラムのセルオブジェクト(None=一致するものがない）
        """
        for cell in sheet_obj.iter_cols(min_row=column_name_line, max_row=column_name_line):
            if cell[0].value is not None and cell[0].value == column_name:
                return cell[0]
        return None

    def get_column_data(self, column_obj, start_line):
        """
        指定したカラムのデータ部を得る
        :param column_obj: 取り出すカラムのオブジェクト
        :param start_line: 取り出し開始位置
        :return: 取り出したCellオブジェクトのリスト
        """
        cell_value_list = []
        for cell in column_obj.parent.iter_rows(min_row=start_line, min_col=column_obj.column,
                                                max_col=column_obj.column):
            if cell[0].value is not None:  # 値が入っていなかったら無視
                cell_value_list.append(cell[0].value)  # 値をリストに追加
        return cell_value_list

    def make_sheet(self, sheet_name_list):
        """
        シートを新規作成する
        :param sheet_name_list:作成するシート名が入れられたリスト
        :return:sheet_list 作成したシートオブジェクトの辞書 { シート名:シートオブジェクト,,, }
        """
        sheet_list = {}
        for sheet_name in sheet_name_list:
            sheet_obj = self.work_book.create_sheet(sheet_name)  # シートを作る
            sheet_list[sheet_name] = sheet_obj  # 辞書に登録

        return sheet_list

    def get_rows_by_searched_column(self, column_obj, keyword, start_line):
        """
        指定したカラムのデータがkeywordである行を抽出してリストにする
        :param column_obj: 検索するカラムのオブジェクト
        :param keyword: 検索するキー
        :param start_line: 検索開始位置
        :return: 抽出した行（セルリスト）のリスト
        """
        cell_list = []
        for cell in column_obj.parent.iter_rows(min_row=start_line):
            if cell[column_obj.column - 1].value is not None and cell[column_obj.column - 1].value == keyword:
                cell_list.append(cell)
        return cell_list

    def get_rows_by_lineNo(self, sheet_obj, start_line, end_line):
        """
        指定した行番号の行のセルをリストにする
        :param sheet_obj: 取得するシートのオブジェクト
        :param start_line: 取得開始行番号
        :param end_line: 取得終了行番号
        :return: 抽出した行（セルリスト）のリスト
        """
        cell_list = []
        for cell in sheet_obj.iter_rows(min_row=start_line, max_row=end_line):
            cell_list.append(cell)
        return cell_list

    def append_rows(self, sheet_obj, rows):
        """
        シートにデータを追加する
        :param sheet_obj:追加するシートのオブジェクト
        :param rows:追加するデータ（セルリストのリスト
        :return:
        """
        for row in rows:
            value_list = []
            for cell in row:
                value_list.append(cell.value)
            sheet_obj.append(value_list)


if __name__ == "__main__":
    name = make_save_filename("C:\\develop\\python\\pythonProject\\divider\\test.data")
    print(name)
    name = make_save_filename("test.data")
    print(name)
    name = make_save_filename("C:\\develop\\python\\pythonProject\\divider\\test")
    print(name)
    name = make_save_filename("C:\\develop\\python\\pythonProject\\divider\\test.py")
    print(name)

    obj = ExcelManager("test1.xlsx")

    sheet = obj.get_sheet("test1")
    print(sheet)
    print(obj.get_sheet("test2"))

    col = obj.get_column(sheet, "品種", 2)
    print(col)
    print(obj.get_column(sheet, "テスト", 2))

    sheet = obj.get_sheet("圃場情報")
    col = obj.get_column(sheet, "品種", 2)
    print(col)
    print(obj.get_column(sheet, "テスト", 2))

    cells = obj.get_column_data(col, 3)
    print(cells)

    print(obj.make_sheet(["test1", "test2", "test3"]))

    cell_line = obj.get_rows_by_searched_column(col, "にこまる", 3)
    print(cell_line)

    obj.close()
